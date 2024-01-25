import csv
from openpyxl import Workbook

# Create a new Excel workbook and a new worksheet
workbook = Workbook()
worksheet = workbook.active

# Define the headers
headers = ["Main website", "Item website", "Item", "Cost", "Currency", "Details", "Details", "Description", "Available"]

# Write the headers to the first row of the worksheet
for column_index, header in enumerate(headers, start=1):
    worksheet.cell(row=1, column=column_index, value=header)

# Open the CSV file
csv_file_path = 'final.csv'
with open(csv_file_path, 'r', newline='', encoding='utf-8') as csvfile:
    reader = csv.reader(csvfile)

    # Iterate over the rows in the CSV file
    for row_index, row in enumerate(reader, start=2):  # start from 2 because 1 is used by headers
        # Write each value from the CSV row to the corresponding cell in the worksheet
        for column_index, value in enumerate(row, start=1):
            worksheet.cell(row=row_index, column=column_index, value=value)

# Save the new Excel file
excel_file_path = 'output.xlsx'
workbook.save(excel_file_path)